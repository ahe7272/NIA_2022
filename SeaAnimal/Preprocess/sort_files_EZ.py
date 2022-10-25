import PySimpleGUI as sg
from glob import glob
import os
import subprocess
import sys
import openpyxl
from datetime import date
import pandas as pd 
import shutil

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')
        integrity = [
                  [sg.Text('SORT DATA', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (10,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Image name list to Excel', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],            
                  [sg.Button('Done', font =("Arial", 15, 'bold'), size=(25, 1))]
                    ]

        window = sg.Window('SORT DATA', integrity, resizable=True, grab_anywhere = True, element_justification='c') 
        return window

def list2excel():
    all_filenames = []
    try:
        for (path, dir, files) in os.walk(values['Path']):
            for jpg in files:
                if jpg[-4:] == '.jpg':
                    all_filenames += [[jpg]]
        listdf = pd.DataFrame.from_records(all_filenames)
        filename = values['Path'] + '/SORT_' + str(date.today()) + '0.xlsx'
        try: 
            if os.path.exists(filename): 
                lastfileno = max([int(x.split(str(date.today()))[-1][0]) for x in glob(filename[:-7]+'*.xlsx')])
                filename = values['Path'] + '/SORT_' + str(date.today()) + str(lastfileno + 1) + '.xlsx'
                listdf.to_excel(filename)
                listlen = len(all_filenames)
            else:
                listdf.to_excel(filename)
                listlen = len(all_filenames)

        except OSError: 
            print("directory를 생성할 수 없습니다.")
            exit()

        return filename, listlen
    except:
        sg.Popup("파일명리스트 엑셀파일 생성에 실패했습니다." + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
    
try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

try:
    flag_cnt = 0
    while True:    
        event, values = window.read()    
        if event == 'Image name list to Excel':
            listexcelpath, listlen = list2excel()

        if event == 'Done':
            sg.Popup("선별을 완료하셨습니다.", font =("Arial", 15), keep_on_top=True)   
            Conch_path = values['Path'] + '/Conch/'
            Hard_path = values['Path'] + '/Hard/'
            Easy_path = values['Path'] + '/Easy/'

            os.makedirs(Conch_path, exist_ok=True)
            os.makedirs(Hard_path, exist_ok=True)
            os.makedirs(Easy_path, exist_ok=True)

            listwb = openpyxl.load_workbook(listexcelpath)
            listws = listwb['Sheet1']
            try: 
                for passROW in range(2, listlen+2):
                    Pure_filename = listws.cell(passROW,2).value[:-4]
                    if listws.cell(passROW,3).value == None:
                        continue
                    elif listws.cell(passROW,3).value.upper() == 'C' :                    
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', Conch_path)
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', Conch_path)
                    elif listws.cell(passROW,3).value.upper() == 'H': 
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', Hard_path)
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', Hard_path)
                    elif listws.cell(passROW,3).value.upper() == 'E': 
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', Easy_path)
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', Easy_path)
                    else:
                        continue
            except:
                sg.Popup("파일 분류에 실패했습니다.", font =("Arial", 15), keep_on_top=True)   
   
        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()