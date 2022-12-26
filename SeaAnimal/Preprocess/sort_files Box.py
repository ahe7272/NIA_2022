import PySimpleGUI as sg
from glob import glob
import os
import subprocess
import sys
import openpyxl
from datetime import date
import pandas as pd 
import shutil

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')
        integrity = [
                  [sg.Text('SORT DATA', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (10,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Image name list to Excel', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],            
                  [sg.Button('Done', font =("Arial", 15, 'bold'), size=(25, 1))]
                    ]

        window = sg.Window('SORT DATA', integrity, resizable=True, grab_anywhere = True, element_justification='c') 
        return window

def list2excel():
    all_filenames = []
    try:
        for (path, dir, files) in os.walk(values['Path']):
            for jpg in files:
                if jpg[-4:] == '.jpg':
                    all_filenames += [[jpg]]
        listdf = pd.DataFrame.from_records(all_filenames)
        filename = values['Path'] + '/SORT_' + str(date.today()) + '0.xlsx'
        try: 
            if os.path.exists(filename): 
                lastfileno = max([int(x.split(str(date.today()))[-1][0]) for x in glob(filename[:-7]+'*.xlsx')])
                filename = values['Path'] + '/SORT_' + str(date.today()) + str(lastfileno + 1) + '.xlsx'
                listdf.to_excel(filename)
                listlen = len(all_filenames)
            else:
                listdf.to_excel(filename)
                listlen = len(all_filenames)

        except OSError: 
            print("directory를 생성할 수 없습니다.")
            exit()

        return filename, listlen
    except:
        sg.Popup("파일명리스트 엑셀파일 생성에 실패했습니다." + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
    
try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

try:
    flag_cnt = 0
    while True:    
        event, values = window.read()    
        if event == 'Image name list to Excel':
            listexcelpath, listlen = list2excel()

        if event == 'Done':
            sg.Popup("선별을 완료하셨습니다.", font =("Arial", 15), keep_on_top=True)   
            H_path = values['Path'] + '/Hemicentrotus/'
            S_path = values['Path'] + '/SeaHare/'
            T_path = values['Path'] + '/Turbo_cornutus/'
            A_path = values['Path'] + '/Asterina_pectinifera/'

            os.makedirs(H_path, exist_ok=True)
            os.makedirs(S_path, exist_ok=True)
            os.makedirs(T_path, exist_ok=True)
            os.makedirs(A_path, exist_ok=True)

            listwb = openpyxl.load_workbook(listexcelpath)
            listws = listwb['Sheet1']
            try: 
                for passROW in range(2, listlen+2):
                    Pure_filename = listws.cell(passROW,2).value[:-4]
                    if listws.cell(passROW,3).value == None:
                        continue
                    elif listws.cell(passROW,3).value.upper() == 'H' :                    
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', H_path + Pure_filename +'.jpg')
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', H_path + Pure_filename +'.json')
                    elif listws.cell(passROW,3).value.upper() == 'S': 
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', S_path + Pure_filename +'.jpg')
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', S_path + Pure_filename +'.json')
                    elif listws.cell(passROW,3).value.upper() == 'T': 
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', T_path + Pure_filename +'.jpg')
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', T_path + Pure_filename +'.json')
                    elif listws.cell(passROW,3).value.upper() == 'A': 
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', A_path + Pure_filename +'.jpg')
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', A_path + Pure_filename +'.json')
                    else:
                        continue
            except:
                sg.Popup("파일 분류에 실패했습니다.", font =("Arial", 15), keep_on_top=True)   
   
        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()