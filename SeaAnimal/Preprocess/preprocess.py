import PySimpleGUI as sg
import os
from glob import glob
import openpyxl
import cv2
from jsonformat import getjsonform
import json

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')

        layout = [
                  [sg.Text('데이터의 경로(폴더)',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Data Folder', font =("Arial", 13), size = (30,1), key='DataPath'), sg.FolderBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('환경정보 데이터 파일(excel 파일)을 선택해주세요.',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Excel File', font =("Arial", 13), size = (30,1), key='WaterinfoPath'), sg.FileBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('비디오 프레임 추출 간격(초)를 입력해주세요.',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Sampling Intervals', font =("Arial", 13),  size = (30,1), key='intervals'), sg.Button('Run',font =("Arial", 13, 'bold'), size=(10,1), key='Run')],
                  [sg.ProgressBar(1, orientation='h', size=(40,20), key='progress')]
                 ]

        window = sg.Window('Preprocessing data', layout, element_justification='c', grab_anywhere = True).Finalize()
        return window

def get_frame(start, end, fps):
    frame_array = [i for i in range(start, end, fps)]
    return frame_array

def update_json(jsonname, imagePath, water_info, h, w):
    objects = getjsonform()
    objects['imagePath'] = imagePath + '.jpg'
    objects['imageHeight'] = h 
    objects['imageWidth'] = w
    objects['Temp'] = water_info['Temp']
    objects['Salinity'] = water_info['Salinity']
    objects['DO'] = water_info['Do']
    objects['pH'] = water_info['pH']
    objects['Transparency'] = water_info['Transparency']
    objects['Longitude'] = water_info['lon']
    objects['Latitude'] = water_info['lat']
    objects['Depth'] = water_info['Depth']

    with open(jsonname + '.json', 'w') as jsonfile:
        json.dump(objects, jsonfile, indent=4)

def clahe_image(img):
    b, g, r = cv2.split(img)
    clahe = cv2.createCLAHE(clipLimit=3.0,tileGridSize=(8,8))
    clahe_b = clahe.apply(b)
    clahe_g = clahe.apply(g)
    clahe_r = clahe.apply(r)
    clahed = cv2.merge((clahe_b, clahe_g, clahe_r))
    return clahed

def preprocess_video(video, interval, savepath, water_info):
    cap = cv2.VideoCapture(video)
    fps = round(cap.get(cv2.CAP_PROP_FPS))
    startframe = int(0 * fps)
    endframe = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) #끝나는시간 * fps
    frame_array = get_frame(startframe, endframe, fps * interval)
    videonamemp4 = os.path.split(video)[-1]
    videoname = os.path.splitext(videonamemp4)[0]
    cnt = 0

    for f in frame_array:
        cnt += 1
        cap.set(cv2.CAP_PROP_POS_FRAMES, f)
        success, image = cap.read()
        
        if success == False:
            print("프레임 추출 실패!")
        else:   
            clahe_img = clahe_image(image)
            h, w, _ = image.shape
            
            cv2.imwrite(savepath + "/" + videoname +'_' + str(cnt) + '.jpg', clahe_img)
            update_json(savepath + "/" + videoname +'_' + str(cnt), videoname +'_' + str(cnt), water_info, h, w)
            
        # print(str(cnt * interval) + "초")
        
        if cv2.waitKey(10) == 27:
            break    

def preprocess_img(image, savepath, water_info):
    imagenamejpg = os.path.split(image)[-1]
    imagename = os.path.splitext(imagenamejpg)[0]
    image = cv2.imread(image)
    clahe_img = clahe_image(image)
    cv2.imwrite(savepath + "/" + imagenamejpg, clahe_img)
    h, w, _ = clahe_img.shape
    update_json(savepath + "/" + imagename, imagename, water_info, h, w)



def get_waterinfo(info_path):
    wb = openpyxl.load_workbook(info_path)
    # 수질 환경정보 excel 파일을 불러와 각 cell의 값을 list로 저장
    ws1 = wb['Sheet1']
    water_info = {}
    for c in range(1, 9):
        if (ws1.cell(2,c).value == None) or (type(ws1.cell(2,c).value) == str):
            return False, water_info
        else:
            water_info[ws1.cell(1,c).value] = ws1.cell(2,c).value
    return True, water_info

# information.xlsx 내 수질정보 열명 확실하게 정하기 및 부경해양에서 작성할 야장 format과 일치
def check_waterinfo(water_info):
    flag_array = [True for i in range(1, 9)]
    if 0 > water_info['Temp'] or water_info['Temp'] > 40:
        flag_array[0] = False
    if 0 > water_info['Salinity'] or water_info['Salinity'] > 40:
        flag_array[1] = False
    if 0 > water_info['Do'] or water_info['Do'] > 15:
        flag_array[2] = False
    if 6 > water_info['pH'] or water_info['pH'] > 9:
        flag_array[3] = False
    if 0 > water_info['Transparency'] or water_info['Transparency'] > 15:
        flag_array[4] = False
    if 33.11 > water_info['lon'] or water_info['lon'] > 38.61:
        flag_array[5] = False
    if 124.6 > water_info['lat'] or water_info['lat'] > 131.87:
        flag_array[6] = False
    if 0 > water_info['Depth']:
        flag_array[7] = False

    return flag_array

m = MakeGUI()
window = m.makegui()

while True:    
    event, values = window.read()     
    progress_bar = window.FindElement('progress')
    if event == 'Run':
        Datapath = values['DataPath']
        info_path = values['WaterinfoPath']
        flag, water_info = get_waterinfo(info_path)
    
        if not flag:
            sg.Popup('수질 환경정보 파일의 값이 숫자가 아니거나 누락되었습니다.', font =("Arial", 13), keep_on_top=True)
            continue

        flag_array = check_waterinfo(water_info)
        if False in flag_array:
            sg.Popup('수질 환경정보 파일에 정상 범위를 벗어난 값이 있습니다.', keep_on_top=True)
            continue

        videolist = glob(Datapath + "/*.mp4")
        imagelist = glob(Datapath + "/*.jpg")
        datalength = len(videolist) + len(imagelist)
        progress_bar.UpdateBar(0, datalength)

        if len(videolist) > 0 :
            interval = int(values['intervals'])
        cnt = 1
        os.makedirs(Datapath + '/processed', exist_ok=True)
        savepath = Datapath + '/processed'
        for video in videolist:
            progress_bar.UpdateBar(cnt, datalength)
            videoname = os.path.split(video)[-1]
            preprocess_video(video, interval, savepath, water_info)
            cnt += 1
   
        for image in imagelist:
            progress_bar.UpdateBar(cnt, datalength)
            imagenamejpg = os.path.split(image)[-1]
            imagename = os.path.splitext(imagenamejpg)[0]
            preprocess_img(image, savepath, water_info)
            cnt += 1
    if event in (None, 'Exit'):
        break

