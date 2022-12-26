import PySimpleGUI as sg
import os
from glob import glob
import openpyxl
import cv2
from jsonformat import getjsonform
import json
import shutil
import datetime
from PIL import Image
import io
import base64
import numpy as np
import datetime

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')

        layout = [
                  [sg.Text('데이터의 경로(폴더)',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Data Folder', font =("Arial", 13), size = (30,1), key='DataPath'), sg.FolderBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('환경정보 데이터 파일(excel 파일)',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Excel File', font =("Arial", 13), size = (30,1), key='WaterinfoPath'), sg.FileBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('비디오 프레임 추출 간격(초).',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Sampling Intervals', font =("Arial", 13),  size = (30,1), key='intervals'), sg.Button('Run',font =("Arial", 13, 'bold'), size=(10,1), key='Run')],
                  [sg.ProgressBar(1, orientation='h', size=(40,20), key='progress')]
                 ]

        window = sg.Window('Preprocessing data', layout, element_justification='c', grab_anywhere = True).Finalize()
        return window

def update_json(jsonname, imagePath, water_info, h, w, source, video_time, frame_no):
    objects = getjsonform()
    objects['imagePath'] = imagePath + '.jpg'
    objects['imageHeight'] = h 
    objects['imageWidth'] = w
    objects['Temperature'] = water_info['Temperature']
    objects['Salinity'] = water_info['Salinity']
    objects['DO'] = water_info['DO']
    objects['pH'] = water_info['pH']
    objects['Longitude'] = water_info['Longitude']
    objects['Latitude'] = water_info['Latitude']
    objects['Depth'] = water_info['Depth']
    objects['Location'] = water_info['Location']
    objects['Weather'] = water_info['Weather']
    objects['Transparency'] = water_info['Transparency']
    objects['Distance'] = water_info['Distance']
    objects['Source_video'] = source
    objects['Video_time'] = video_time
    objects['Frame_no'] = frame_no
    objects['Date_created'] = str(water_info['Date_created'])[:10]
    if len(objects['Source_video'].split('ROV')) > 1:
        objects['Collection_method'] = 'ROV'
    else:
        objects['Collection_method'] = 'Diver'

    with open(jsonname + '.json', 'w') as jsonfile:
        json.dump(objects, jsonfile, indent=4)

def preprocess_img(image, savepath, water_info):
    imagenamejpg = os.path.split(image)[-1]
    imagename = os.path.splitext(imagenamejpg)[0]
    image = cv2.imread(image)
    h, w, _ = image.shape
    if (h == 3840) and  (w == 2160):
        image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
        h = 2160
        w = 3840
    elif (h not in [2160, 3840]) or (w not in [3840, 2160]):
        sg.Popup(imagename + '이미지 비율이 안 맞습니다.'+str(h) + ' ' + str(w), font =("Arial", 15), keep_on_top=True)
        return False, water_info 
    source = '_'.join(imagename.split('_')[:-1]) +'.mp4'
    videotime = str(datetime.timedelta(seconds= ((int(imagename.split('_')[-1].split('.')[0]))-1)))
    frameno = (int(imagename.split('_')[-1].split('.')[0])-1)*60

    cv2.imwrite(savepath + "/" + imagenamejpg, image)
    update_json(savepath + "/" + imagename, imagename, water_info, h, w, source, videotime, frameno)

def get_waterinfo(info_path):
    water_info_cols = ['Temperature', 'Salinity', 'DO', 'pH', 'Longitude', 'Latitude', 'Depth', 'Location' , 'Weather', 'Transparency', 'Date_created', 'Distance']
    try:
        wb = openpyxl.load_workbook(info_path)
        # 수질 환경정보 excel 파일을 불러와 각 cell의 값을 list로 저장
        ws1 = wb['Sheet1']
    except:
        sg.Popup('excel파일이 경로에 없거나 Sheet 네임이 \n "Sheet1"과 "Sheet2"로 정확하게 입력되었는지 확인해주세요.', font =("Arial", 15), keep_on_top=True)
    water_info = {}
    meta_imgs = []
    for c1 in range(1, 13):
        if c1 == 12:
            if ws1.cell(2,c1).value == None:
                water_info['Distance'] = 0
            else:
                water_info[ws1.cell(1,c1).value] = ws1.cell(2,c1).value
            continue
        if ws1.cell(1, c1).value not in water_info_cols:
            sg.Popup('환경정보 파일의 Sheet1 열명이 Format에 맞지 않습니다. \nExcel 파일을 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
            return False, water_info, meta_imgs
        if c1 == 8:
            if (ws1.cell(2,c1).value == None) or (type(ws1.cell(2,c1).value) != str):
                sg.Popup('Excel Sheet1의 Location 값을 확인해주세요.', font =("Arial", 15), keep_on_top=True)
                return False, water_info, meta_imgs
            water_info[ws1.cell(1,c1).value] = ws1.cell(2,c1).value
            continue
        if c1 == 11:
            if (ws1.cell(2,c1).value == None) or (type(ws1.cell(2,c1).value) != datetime.datetime):
                sg.Popup('Excel Sheet1의 Date_created 값을 확인해주세요.', font =("Arial", 15), keep_on_top=True)
                return False, water_info, meta_imgs
            water_info[ws1.cell(1,c1).value] = ws1.cell(2,c1).value
            continue
        
        if (ws1.cell(2,c1).value == None) or (type(ws1.cell(2,c1).value) == str):
            sg.Popup('환경정보 값이 누락되었거나 오탈자가 들어 있습니다.\nExcel 파일을 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
            return False, water_info, meta_imgs
        else:
            water_info[ws1.cell(1,c1).value] = ws1.cell(2,c1).value

    return True, water_info, meta_imgs

def check_waterinfo(water_info):
    flag_array = [True for i in range(1, 12)]
    if 0 > water_info['Temperature'] or water_info['Temperature'] > 40:
        flag_array[0] = False
    if 0 > water_info['Salinity'] or water_info['Salinity'] > 40:
        flag_array[1] = False
    if 0 > water_info['DO'] or water_info['DO'] > 15:
        flag_array[2] = False
    if 6 > water_info['pH'] or water_info['pH'] > 10:
        flag_array[3] = False
    if 33.11 > water_info['Latitude'] or water_info['Latitude'] > 38.61:
        flag_array[5] = False
    if 124.6 > water_info['Longitude'] or water_info['Longitude'] > 132:
        flag_array[6] = False
    if 0 > water_info['Depth']:
        flag_array[7] = False
    if water_info['Weather'] not in [1,2,3]:
        flag_array[8] = False
    if 0 > water_info['Transparency'] or water_info['Transparency'] > 20:
        flag_array[8] = False
    created_date = str(water_info['Date_created'])
    if (int(created_date[:4]) != 2022):
        flag_array[9] = False
    if (int(created_date[5:7]) > 12) or (int(created_date[5:7]) < 6):
        flag_array[9] = False
    if water_info['Distance'] not in [0, 0.5, 1.0, 1.5]:
        flag_array[10] = False
    return flag_array

m = MakeGUI()
window = m.makegui()

while True:    
    event, values = window.read()     
    progress_bar = window['progress']
    if event == 'Run':
        Datapath = values['DataPath']
        info_path = values['WaterinfoPath']
        flag, water_info, meta_imgs = get_waterinfo(info_path)
        info_name = os.path.split(info_path)[-1]
        if not flag:
            continue

        flag_array = check_waterinfo(water_info)
        if False in flag_array:
            sg.Popup('수질 환경정보 파일에 정상 범위를 벗어난 값이 있습니다.', keep_on_top=True)
            continue

        videolist = glob(Datapath + "/*.mp4")
        imagelist = glob(Datapath + "/*.jpg")
        datalength = len(videolist) + len(imagelist)
        progress_bar.UpdateBar(0, datalength)
        if len(videolist) > 0 :
            interval = int(values['intervals'])
        cnt = 1
        savepath = os.path.split(Datapath)[0] + '/' + str(os.path.split(Datapath)[1]) + '_Preprocessed'
        os.makedirs(savepath, exist_ok=True)
        
        for image in imagelist:
            progress_bar.UpdateBar(cnt, datalength)
            imagenamejpg = os.path.split(image)[-1]
            imagename = os.path.splitext(imagenamejpg)[0]
            preprocess_img(image, savepath, water_info)
            cnt += 1

        if cnt == datalength +1 :
            sg.Popup('정제 완료^^!', font =("Arial", 13), keep_on_top=True)
            break

        
    if event in (None, 'Exit'):
        break

