import os
import json
import numpy as np
import glob
import openpyxl
import shutil
from datetime import date

def classname_check(objects):
    classname_error = ""
    total_obj = len(objects['shapes'])
    classnameflag = True
    for t in range(total_obj):
        if objects['shapes'][t]['label'] not in classname:
            classname_error += 'Annotation 이름 에러!\n내용: ' + objects['shapes'][t]['label'] + '\n'
            classnameflag = False
    return classnameflag, classname_error

def attribute_value(objects):
    attr_error = ""
    if len(objects) == 19:
        return True, attr_error
    else:
        attr_error += '속성 개수 에러!\n' + str(len(objects)) + ' 개로 속성값에 이상이 있습니다.' + '\n'
        return False, attr_error

def size(objects, minsize):
    total_obj = len(objects['shapes'])
    for t in range(total_obj - 1, -1, -1):
        points = objects['shapes'][t]['points'] 
        points = np.array(points, np.int32)

        # boundingbox
        if objects['shapes'][t]['shape_type'] == 'rectangle':
            lefttopx, lefttopy = points[0]
            rightdownx, rightdowny = points[1]

        #polygon
        else:
            x = points[:, 0]
            y = points[:, 1]
            lefttopx = min(x)
            rightdownx = max(x)
            lefttopy = min(y)
            rightdowny = max(y)
        w = max(lefttopx, rightdownx) - min(lefttopx, rightdownx)
        h = max(lefttopy, rightdowny) - min(lefttopy, rightdowny)
        area = w * h 
        if area <= minsize:
            del(objects['shapes'][t])   
    return objects
    
def savejson(objects, jsonfile):
    with open(jsonfile, 'w') as Jsonfile:
        json.dump(objects, Jsonfile, indent=4)

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
    return objects

def label_exist(objects):
    label_error = ""
    labelflag = True
    if len(objects['shapes']) == 0:
        label_error += '사진에 Annotation이 존재하지 않습니다. 작업을 완료하시고 무결성 검사를 진행해 주세요.\n'
        labelflag = False
    return labelflag, label_error

classname =['Asterias_amurensis', 'Asterina_pectinifera', 'Conch', 'Ecklonia_cava', 'Heliocidaris_crassispina','Hemicentrotus','Sargassum',  'Sea_hare', 'Turbo_cornutus']

def metacheck(path):
    errorlist = ""
    flag = True
    try:
        wb = openpyxl.load_workbook(path[0]) 
        ws2 = wb['Sheet2']
        for row in range(2, ws2.max_row+1):
            try:
                float(ws2.cell(row, 3).value)
                if ws2.cell(row, 3).value > 500 or ws2.cell(row, 3).value < 0: 
                    errorlist += 'Size 에러!\nExcel 파일 내 Size열 ' +str(row) + ' 행에 숫자가 정상범위를 벗어납니다. 확인 및 수정 후 Metadata 작업을 진행해 주세요.\n'
                    flag = False
            except ValueError:
                errorlist += 'Size 에러!\nExcel 파일 내 Size열 ' +str(row) + ' 행의 값이 숫자가 아닙니다. 확인 및 수정 후 Metadata 작업을 진행해 주세요.\n'
                flag = False            
            try:
                float(ws2.cell(row, 4).value) 
                if ws2.cell(row, 4).value >1000 or ws2.cell(row, 4).value < 0: 
                    errorlist += 'Weight 에러!\nExcel 파일 내 Weight열 ' +str(row) + ' 행에 숫자가 정상범위를 벗어납니다. 확인 및 수정 후 Metadata 작업을 진행해 주세요.\n'
                    flag = False
            except ValueError:
                errorlist += 'Weight 에러!\nExcel 파일 내 Weight열 ' +str(row) + ' 행의 값이 숫자가 아닙니다. 확인 및 수정 후 Metadata 작업을 진행해 주세요.\n'
                flag = False
        return flag, errorlist
    except:
        return flag, errorlist
        
def check(jsonfile, minsize, path):
    errorlist = ""
    jsonpath = path + '/' + jsonfile
    imagefile = os.path.splitext(jsonfile)[0] + '.jpg'
    imagepath = path + '/' + imagefile
    processed_path = path + '/' + str(date.today()) +'/'

    excelpath = path + '/*.xlsx'
    os.makedirs(processed_path, exist_ok=True)

    objects = getjson(jsonpath)
    objects = size(objects, minsize)
    
    savejson(objects, jsonpath)
    classnameflag, classname_error = classname_check(objects)
    if classnameflag:
        classes = True
    else:
        classes = False
        errorlist += str(imagefile) +' 파일 - ' + classname_error + '\n'

    attributeflag, attribute_error = attribute_value(objects)
    if attributeflag:
        attribute = True
    else:
        attribute = False
        errorlist += str(imagefile) + ' 파일 - ' + attribute_error + '\n'
    
    metaflag, meta_error = metacheck(glob.glob(excelpath))
    if metaflag:
        meta = True
    else:
        meta = False

    labelflag, label_error = label_exist(objects)
    if labelflag:
        labels = True
    else:
        labels = False
        errorlist += str(imagefile) + ' 파일 - ' + label_error + '\n'

    if classes and attribute and labels and meta:
        shutil.move(jsonpath, processed_path + jsonfile)
        shutil.move(imagepath, processed_path + imagefile)
    return errorlist 
    


