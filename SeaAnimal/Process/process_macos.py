# 에러출력되게 변경, error파일 생성-> 완료된 파일만 done파일에 들어가게 
import PySimpleGUI as sg
import base64
import pyperclip
from glob import glob
import os
from integrity_check import check, metacheck
import subprocess
import sys
import json
import openpyxl

class MakeGUI():
    def getimage(self, image):
        with open(image, 'rb') as img:
            base64_string = base64.b64encode(img.read())    
        return base64_string

    def makegui(self):
        sg.theme('DarkAmber')
        Step1 = [
                  [sg.Text('ANNOTATION', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text(' PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (22,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(12, 1), key='Select_Folder')],
                  [sg.Text('  ID  ', font =("Arial", 15)), sg.InputText(key='ID', font =("Arial", 15), size = (23,1)), sg.Button('Labelme', font =("Arial", 15, 'bold'), size=(12, 1))],
                  [sg.Text(' Asterias Amurensis ', font =("Arial", 12)), sg.Text('  Asterina Pectinifera  ', font =("Arial", 12)), sg.Text('           Conch',font =("Arial", 12))], 
                  [
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Asterias Amurensis.png'), key='Asterias Amurensis'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Asterina Pectinifera.png'), key='Asterina Pectinifera'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Conch.png'), key='Conch')
                  ],
                  [sg.Text('      EckloniaCava    ', font =("Arial", 12)), sg.Text('Heliocidaris Crassispina', font =("Arial", 12)), sg.Text('   Hemicentrotus', font =("Arial", 12))],
                  [
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/EckloniaCava.png'), key='EckloniaCava'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Heliocidaris Crassispina.png'), key='Heliocidaris Crassispina'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Hemicentrotus.png'), key='Hemicentrotus')
                  ],
                  [sg.Text('        Sargassum        ', font =("Arial", 12)), sg.Text('            SeaHare            ', font =("Arial", 12)), sg.Text('   Turbo Cornutus', font =("Arial", 12))],
                  [
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Sargassum.png'), key='Sargassum'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/SeaHare.png'), key='SeaHare'),
                    sg.Button('', image_data=self.getimage('C:/Users/Administrator/Documents/Github/NIA_2022/SeaAnimal/Process/example_image/Turbo Cornutus.png'), key='Turbo Cornutus')
                  ]
                ]
        Step2 = [
                  [sg.Text('FILL IN METADATA', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('라벨한 순서대로 크기와 무게가 입력됩니다.', font =("Arial", 15, 'bold'))],
                  [sg.Button('입력', font=("Arial", 15, 'bold'), size=(39, 5), key='Meta')],
                  [sg.Text('', font =("Arial", 10))],
                  [sg.Text('INTEGRITY CHECK', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('에러 메세지를 확인하여 해당 파일을 수정해 주세요.', font =("Arial", 15, 'bold'))],
                  [sg.Text(' - 라벨명 에러: Labelme 재실행 및 annotation 이름 수정', font =("Arial", 13))],
                  [sg.Text(' - 속성 개수 에러: ID 입력 후 Labelme 재실행', font =("Arial", 13))],
                  [sg.Button('무결성 검사', font=("Arial", 15, 'bold'), size=(39, 5), key='integrity')], 
                  [sg.Text('', font =("Arial", 10))],
                  [sg.Button('Exit', font=("Arial", 15, 'bold'),size=(39, 5))]
                ]
        tab_group = [
                     [sg.TabGroup(
                            [[
                            sg.Tab('STEP1', Step1),
                            sg.Tab('STEP2', Step2)
                            ]], 
                            tab_location='centertop',
                            font = ("Arial", 20),
                            tab_border_width = 2,               
                            selected_background_color='White',
                            selected_title_color='Black',
                            border_width=10)]        
                    ]
        window = sg.Window('Processing Data', tab_group, resizable=True, grab_anywhere = True, element_justification='c') 
        return window
        
def runLabelme():
    subprocess.Popen(subprocess.getoutput('type labelme')[11:])

try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

object = ['Asterias Amurensis', 'Asterina Pectinifera', 'Conch', 'EckloniaCava', 'Heliocidaris Crassispina','Hemicentrotus','Sargassum',  'SeaHare', 'Turbo Cornutus']
minsize = 1024

try:
    while True:    
        event, values = window.read()    
        Path = values['Path']
        ID = values['ID']
        if event in object:
            pyperclip.copy(event)

        if event == 'Labelme':
            imagelist = glob(values['Path'] + '/*.jpg')
            for image in imagelist:
                jpgfile = os.path.split(image)[-1]
                imagename = os.path.splitext(jpgfile)[0]
                jsonfile = Path + '/' + imagename + '.json'
                imagefile = Path + '/' + imagename + '.jpg'
            runLabelme()

        if event == 'Meta':
            jsonlist = glob(values['Path'] + '/*.json')
            wb = openpyxl.load_workbook(glob(values['Path'] + '/*.xlsx')[0]) 
            ws2 = wb['Sheet2']
            for each_json in jsonlist:
                jsonfullname = os.path.split(each_json)[-1]
                jsonname = os.path.splitext(jsonfullname)[0]
                with open(each_json) as jsontosave:
                    objects = json.load(jsontosave)
                objects['ID'] = ID
                labelnames = []
                for each_label in objects['shapes']:
                    each_label['Size'] = 0
                    each_label['Weight'] = 0
                    labelnames.append(each_label['label'])
                object_list = []
                for row in range(2, ws2.max_row+1):
                    if jsonname == ws2.cell(row, 1).value:
                        object_list.append([row, ws2.cell(row, 3).value, ws2.cell(row, 4).value, ws2.cell(row, 5).value])
                        
                while len(object_list) > 0:
                    obj = object_list.pop(0)
                    if obj[1] not in object:
                        sg.Popup("Excel 파일에 "+ jsonname+"에 해당하는 클래스 이름에 오타가 있습니다. \n내용: "+ ws2.cell(obj[0], 3).value + "\nExcel 파일 내에 Class 이름을 수정하고 다시 Metadata 입력 작업을 해주세요.", font =("Arial", 15), keep_on_top=True)
                        obj = object_list.pop(0)
                    if obj[1] in labelnames:
                        label_indices = [i for i, x in enumerate(labelnames) if x == obj[1]]
                        for i in label_indices:
                            if objects['shapes'][i]['Size'] == 0 and objects['shapes'][i]['Weight'] == 0:
                                objects['shapes'][i]['Size']  = obj[2]
                                objects['shapes'][i]['Weight']  = obj[3]
                                break
                            else:
                                continue
                with open(each_json, 'w') as jsontosave:
                    json.dump(objects, jsontosave, indent=4)
            sg.Popup("Metadata 입력이 완료되었습니다. 무결성 검사를 진행해 주세요.", font =("Arial", 15), keep_on_top=True)    

        if event == 'integrity':
            metaerrors, metaflag = metacheck(glob(values['Path'] + '/*.xlsx')[0])            
            jsonlist = glob(values['Path'] + '/*.json')
            jsonerrors = ""
            for Json in jsonlist:
                jsonfile = os.path.split(Json)[-1]
                jsonerrors += check(jsonfile, minsize, Path)
            if len(jsonerrors) > 0:
                sg.Popup(jsonerrors, font =("Arial", 15), keep_on_top=True)
            if metaflag == False:
                sg.Popup(metaerrors, font =("Arial", 15), keep_on_top=True)
            else:
                sg.Popup("작업완료된 파일이 Done 폴더에 저장되었습니다. \nExit을 누르고 NAS에 업로드를 시작해 주세요.", font =("Arial", 15), keep_on_top=True)

        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()


