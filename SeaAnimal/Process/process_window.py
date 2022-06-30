import PySimpleGUI as sg
import pyperclip
from glob import glob
import os
from integrity_check import check, metacheck
import subprocess
import sys
import json
import openpyxl
import numpy as np
import requests
from io import BytesIO
import pickle
from PIL.ExifTags import TAGS
from PIL import Image
from glob import glob
import piexif


# 전체 공유 주소: https://ifh.cc/v-Xthdhr.WYZ2DV.58vJQw.r9g0lJ.xx07bk.G9dfGR.zQpXO1.m2Rpvt.PSZRNZ

Asterias_amurensis = BytesIO(requests.get('https://ifh.cc/g/Xthdhr.png').content).read()
Asterina_pectinifera = BytesIO(requests.get('https://ifh.cc/g/WYZ2DV.png').content).read()
Conch = BytesIO(requests.get('https://ifh.cc/g/58vJQw.png').content).read()
Ecklonia_cava = BytesIO(requests.get('https://ifh.cc/g/r9g0lJ.png').content).read()
Heliocidaris_crassispina = BytesIO(requests.get('https://ifh.cc/g/xx07bk.png').content).read()
Hemicentrotus = BytesIO(requests.get('https://ifh.cc/g/G9dfGR.png').content).read()
Sargassum = BytesIO(requests.get('https://ifh.cc/g/zQpXO1.png').content).read()
Sea_hare = BytesIO(requests.get('https://ifh.cc/g/m2Rpvt.png').content).read()
Turbo_cornutus = BytesIO(requests.get('https://ifh.cc/g/PSZRNZ.png').content).read()

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')
        Step1 = [
                  [sg.Text('ANNOTATION', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text(' PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (22,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(12, 1), key='Select_Folder')],
                  [sg.Text('  ID  ', font =("Arial", 15)), sg.InputText(key='userID', font =("Arial", 15), size = (23,1)), sg.Button('Labelme', font =("Arial", 15, 'bold'), size=(12, 1))],
                  [sg.Text(' Asterias_amurensis ', font =("Arial", 12)), sg.Text('  Asterina_pectinifera  ', font =("Arial", 12)), sg.Text('           Conch',font =("Arial", 12))], 
                  [
                    sg.Button('', image_data = Asterias_amurensis,  key='Asterias_amurensis'),
                    sg.Button('', image_data = Asterina_pectinifera, key='Asterina_pectinifera'),
                    sg.Button('', image_data = Conch, key='Conch')
                  ],
                  [sg.Text('      Ecklonia_cava    ', font =("Arial", 12)), sg.Text('Heliocidaris_crassispina', font =("Arial", 12)), sg.Text('   Hemicentrotus', font =("Arial", 12))],
                  [
                    sg.Button('', image_data = Ecklonia_cava, key='Ecklonia_cava'),
                    sg.Button('', image_data = Heliocidaris_crassispina, key='Heliocidaris_crassispina'),
                    sg.Button('', image_data = Hemicentrotus, key='Hemicentrotus')
                  ],
                  [sg.Text('        Sargassum        ', font =("Arial", 12)), sg.Text('            Sea_hare            ', font =("Arial", 12)), sg.Text('   Turbo_cornutus', font =("Arial", 12))],
                  [
                    sg.Button('', image_data = Sargassum, key='Sargassum'),
                    sg.Button('', image_data = Sea_hare, key='Sea_hare'),
                    sg.Button('', image_data = Turbo_cornutus, key='Turbo_cornutus')
                  ]
                ]
        Step2 = [
                  [sg.Text('FILL IN METADATA', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('라벨한 순서대로 크기와 무게가 입력됩니다.', font =("Arial", 15, 'bold'))],
                  [sg.Text(' - Excel 파일 내 Class 이름 에러: Excel 파일 내 Class 이름 수정', font =("Arial", 13))],
                  [sg.Text(' - Size 에러: Excel 파일 내 Size 값 수정', font =("Arial", 13))],
                  [sg.Text(' - Weight 에러: Excel 파일 내 Weight 값 수정', font =("Arial", 13))],
                  [sg.Button('입력', font=("Arial", 15, 'bold'), size=(39, 3), key='Meta')],
                  [sg.Text('', font =("Arial", 10))],
                  [sg.Text('INTEGRITY CHECK', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('에러 메세지를 확인하여 해당 파일을 수정해 주세요.', font =("Arial", 15, 'bold'))],
                  [sg.Text(' - Annotation 이름 에러: Labelme 재실행 및 annotation 이름 수정', font =("Arial", 13))],
                  [sg.Text(' - 속성 개수 에러: json 파일 확인 및 담당 PM과 소통', font =("Arial", 13))],
                  [sg.Text(' - ID 에러: STEP1에 ID 입력 후 다시 Metadata 입력', font =("Arial", 13))],
                  [sg.Button('무결성 검사', font=("Arial", 15, 'bold'), size=(39, 3), key='integrity')], 
                  [sg.Text('', font =("Arial", 10))],
                  [sg.Button('Exit', font=("Arial", 15, 'bold'),size=(39, 3))]
                ]
        tab_group = [
                     [sg.TabGroup(
                            [[
                            sg.Tab('STEP1', Step1),
                            sg.Tab('STEP2', Step2)
                            ]], 
                            tab_location='centertop',
                            font = ("Arial", 20),
                            tab_border_width = 2,               
                            selected_background_color='White',
                            selected_title_color='Black',
                            border_width=10)]        
                    ]
        window = sg.Window('Processing Data', tab_group, resizable=True, grab_anywhere = True, element_justification='c') 
        return window

def changeexif(imagefile, ID):
    image = Image.open(imagefile)
    exifData = image._getexif()
    
    if exifData is None:
        exifData = {}

    data = pickle.dumps(ID)
    exif_ifd = {piexif.ExifIFD.UserComment: data}
    zeroth_ifd = {piexif.ImageIFD.Artist: ID.encode('ascii')}
    exif_dict = {"0th" : zeroth_ifd, "Exif": exif_ifd}

    exif_dat = piexif.dump(exif_dict)
    image.save(imagefile, exif=exif_dat)


def runLabelme():
    conda_dirs = ["c:/Users/users/anaconda3", "d:/Users/users/anaconda3", 
                  "c:/Users/Admin/anaconda3", "d:/Users/Admin/anaconda3", "c:/Users/admin/anaconda3", "d:/Users/admin/anaconda3", 
                  "c:/Users/Administrator/anaconda3", "d:/Users/Administrator/anaconda3",
                  "c:/Programdata/anaconda3", "d:/Programdata/anaconda3", 
                  "c:/Users/{}/Anaconda3".format(os.getlogin()), "d:/Users/{}/Anaconda3".format(os.getlogin()), 
                  "c:/사용자/{}/Anaconda3".format(os.getlogin()), "d:/사용자/{}/Anaconda3".format(os.getlogin()),
                  "c:/anaconda3", "c:/Anaconda3", "d:/anaconda3", "d:/Anaconda3"]
    try:
        for Dir in conda_dirs:
            if os.path.isdir(Dir):
                python_file = Dir + "/envs/nia/pythonw.exe"
                labelme_file =  python_file[:-11] + "/Lib/site-packages/labelme/__main__.py"

            if Dir == "no dir":
                sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
        
        if np.__version__[:4] == '1.22':
        
            subprocess.Popen(python_file + ' ' + labelme_file,  stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)

    except:
        sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)

try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

object = ['Asterias_amurensis', 'Asterina_pectinifera', 'Conch', 'Ecklonia_cava', 'Heliocidaris_crassispina','Hemicentrotus','Sargassum',  'Sea_hare', 'Turbo_cornutus']
minsize = 1024

try:
    while True:    
        event, values = window.read()    
        Path = values['Path']
        userID = values['userID']
        if event in object:
            pyperclip.copy(event)

        if event == 'Labelme':
            imagelist = glob(values['Path'] + '/*.jpg')
            for image in imagelist:
                jpgfile = os.path.split(image)[-1]
                imagename = os.path.splitext(jpgfile)[0]
                jsonfile = Path + '/' + imagename + '.json'
                imagefile = Path + '/' + imagename + '.jpg'
                changeexif(image, userID)
            runLabelme()

        if event == 'Meta':
            jsonlist = glob(values['Path'] + '/*.json')
            meta_success = True
            try: 
                flag, meta_error = metacheck(glob(values['Path'] + '/*.xlsx'))
                if len(meta_error) >0 :
                    sg.Popup(meta_error, font =("Arial", 15), keep_on_top=True)
                    continue
                wb = openpyxl.load_workbook(glob(values['Path'] + '/*.xlsx')[0]) 
                ws2 = wb['Sheet2']
                for each_json in jsonlist:
                    jsonfullname = os.path.split(each_json)[-1]
                    jsonname = os.path.splitext(jsonfullname)[0]
                    with open(each_json) as jsontosave:
                        objects = json.load(jsontosave)
                        objects['ID'] = userID
                    labelnames = []
                    for each_label in objects['shapes']:
                        each_label['Size'] = 0
                        each_label['Weight'] = 0
                        labelnames.append(each_label['label'])
                    object_list = []
                    for row in range(2, ws2.max_row+1):
                        if jsonname == ws2.cell(row, 1).value:
                            object_list.append([row, ws2.cell(row, 2).value, ws2.cell(row, 3).value, ws2.cell(row, 4).value])
                            
                    while len(object_list) > 0:
                        obj = object_list.pop(0)
                        if obj[1] not in object:
                            sg.Popup("Excel 파일 내 Class 이름 에러!\nExcel 파일에 "+ jsonname+"에 해당하는 Class 이름에 오타가 있습니다. \n내용: "+ ws2.cell(obj[0], 2).value + "\nExcel 파일 내에 Class 이름을 수정하고 다시 Metadata 입력 작업을 해주세요.\n", font =("Arial", 15), keep_on_top=True)
                            meta_success = False
                            continue
                        if obj[1] in labelnames:
                            label_indices = [i for i, x in enumerate(labelnames) if x == obj[1]]
                            for i in label_indices:
                                if objects['shapes'][i]['Size'] == 0 and objects['shapes'][i]['Weight'] == 0:
                                    objects['shapes'][i]['Size']  = obj[2]
                                    objects['shapes'][i]['Weight']  = obj[3]
                                    break
                                else:
                                    continue
                    with open(each_json, 'w') as jsontosave:
                        json.dump(objects, jsontosave, indent=4)
            except:
                for each_json in jsonlist:
                    jsonfullname = os.path.split(each_json)[-1]
                    jsonname = os.path.splitext(jsonfullname)[0]
                    with open(each_json) as jsontosave:
                        objects = json.load(jsontosave)
                        objects['ID'] = userID
                    for each_label in objects['shapes']:
                        each_label['Size'] = 0
                        each_label['Weight'] = 0
                    with open(each_json, 'w') as jsontosave:
                        json.dump(objects, jsontosave, indent=4)

            if meta_success:
                sg.Popup("Metadata 입력이 완료되었습니다. 무결성 검사를 진행해 주세요.\n", font =("Arial", 15), keep_on_top=True)    

        if event == 'integrity':
            jsonlist = glob(values['Path'] + '/*.json')
            jsonerrors = ""
            metaerrors = ""
            if values['userID'] == "":
                jsonerrors += 'ID 에러!\nID를 정확하게 입력하고 Metadata 입력을 다시 진행해 주세요.\n'
            for Json in jsonlist:
                jsonfile = os.path.split(Json)[-1]
                jsonerrors += check(jsonfile, minsize, Path)
            if len(jsonerrors) > 0:
                sg.Popup(jsonerrors, font =("Arial", 15), keep_on_top=True)
            else:
                sg.Popup("작업완료된 파일이 Processed 폴더에 저장되었습니다. \n 작업 폴더에 남은 파일이 없는지 확인해주세요.", font =("Arial", 15), keep_on_top=True)
        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()