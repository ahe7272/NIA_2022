import PySimpleGUI as sg
import os
from glob import glob
import openpyxl
import cv2
from jsonformat import getjsonform
import json
import copy
from skimage.metrics import structural_similarity as ssim
from PIL import Image
import io
import base64
import numpy as np
import datetime

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')

        layout = [
                  [sg.Text('데이터의 경로(폴더)',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Data Folder', font =("Arial", 13), size = (30,1), key='DataPath'), sg.FolderBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('환경정보 데이터 파일(excel 파일)을 선택해주세요.',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Excel File', font =("Arial", 13), size = (30,1), key='WaterinfoPath'), sg.FileBrowse('SELECT', font =("Arial", 13, 'bold'), size=(10, 1))],
                  [sg.Text('비디오 프레임 추출 간격(초)를 입력해주세요.',font =("Arial", 13, 'bold'), size=(40, 1))],
                  [sg.InputText('Sampling Intervals', font =("Arial", 13),  size = (30,1), key='intervals'), sg.Button('Run',font =("Arial", 13, 'bold'), size=(10,1), key='Run')],
                #   [sg.Text('프레임 유사도 한계값(0.0~1.0)을 설정해 주세요.',font =("Arial", 13, 'bold'), size=(40, 1))],
                #   [sg.InputText('Similarity Threshold', font =("Arial", 13),  size = (30,1), key='thres'), sg.Button('Run',font =("Arial", 13, 'bold'), size=(10,1), key='Run')],
                  [sg.ProgressBar(1, orientation='h', size=(40,20), key='progress')]
                 ]

        window = sg.Window('Preprocessing data', layout, element_justification='c', grab_anywhere = True).Finalize()
        return window

def get_frame(start, end, fps):
    frame_array = [i for i in range(start, end, fps)]
    return frame_array

def img_to_b64(imgfile):
    if isinstance(imgfile, (np.ndarray)) == False:
        imgfile = cv2.imread(imgfile)
    imgfile = Image.fromarray(imgfile)
    f = io.BytesIO()
    imgfile.save(f, format='PNG')
    img_bin = f.getvalue()
    if hasattr(base64, 'encodebytes'):
        img_b64 = base64.encodebytes(img_bin)
    else:
        img_b64 = base64.encodestring(img_bin)
    return img_b64

def update_json(jsonname, imagePath, water_info, h, w, Databin, source, video_time, frame_no):
    objects = getjsonform()
    objects['imagePath'] = imagePath + '.jpg'
    objects['imageHeight'] = h 
    objects['imageWidth'] = w
    objects['Transparency'] = water_info['Transparency']
    objects['Latitude'] = water_info['Latitude']
    objects['Longitude'] = water_info['Longitude']
    objects['Site_Type'] = water_info['Site_Type']
    objects['Depth'] = water_info['Depth']
    objects['Source_video'] = source
    objects['Video_time'] = video_time
    objects['Frame_no'] = frame_no
    objects['Origin_img'] = Databin.decode('utf8')
    with open(jsonname + '.json', 'w') as jsonfile:
        json.dump(objects, jsonfile, indent=4)

def clahe_image(img):
    b, g, r = cv2.split(img)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    clahe_b = clahe.apply(b)
    clahe_b = cv2.normalize(clahe_b, None, 0, 255, cv2.NORM_MINMAX)
    clahe_g = clahe.apply(g)
    clahe_g = cv2.normalize(clahe_g, None, 0, 255, cv2.NORM_MINMAX)
    clahe_r = clahe.apply(r)
    clahe_r = cv2.normalize(clahe_r, None, 0, 255, cv2.NORM_MINMAX)
    clahed = cv2.merge((clahe_b, clahe_g, clahe_r))
    return clahed

def getSimilarity(img1, img2):
    gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

    crop1 = gray1[int(gray1.shape[0] * 0.25):int(gray1.shape[0] * 0.75), int(gray1.shape[1] * 0.25):int(gray1.shape[1] * 0.75)]
    crop2 = gray2[int(gray1.shape[0] * 0.25):int(gray1.shape[0] * 0.75), int(gray1.shape[1] * 0.25):int(gray1.shape[1] * 0.75)]

    (score, _) = ssim(crop1, crop2, full=True)
    # (score, _) = ssim(gray1, gray2, full=True)

    return score

def preprocess_video(video, interval, savepath, water_info):
    cap = cv2.VideoCapture(video)
    fps = round(cap.get(cv2.CAP_PROP_FPS))
    startframe = int(0 * fps)
    endframe = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) #끝나는시간 * fps
    frame_array = get_frame(startframe, endframe, fps * interval)
    videonamemp4 = os.path.split(video)[-1]
    videoname = os.path.splitext(videonamemp4)[0]
    cnt = 0
    # threshold = float(values['thres'])
    threshold = 1

    nFrame = 1
    Memory = None
    for f in frame_array:
        cnt += 1
        cap.set(cv2.CAP_PROP_POS_FRAMES, f)
        success, image = cap.read()
        if success == False:
            print("프레임 추출 실패!")
        else:                
            if nFrame == 1:
                Memory = copy.deepcopy(image)
                nFrame += 1
                Databin = img_to_b64(image)
                clahe_img = clahe_image(image)
                h, w, _ = image.shape
                if (h == 3840) and  (w == 2160):
                    clahe_img = cv2.rotate(clahe_img, cv2.ROTATE_90_CLOCKWISE)
                    h = 2160
                    w = 3840
                elif (h not in [2160, 3840]) or (w not in [3840, 2160]):
                    sg.Popup(videoname + '비디오 비율이 안 맞습니다.'+str(h) + ' ' + str(w), font =("Arial", 15), keep_on_top=True)
                    return False, water_info 
                video_time = str(datetime.timedelta(seconds= f/60))
                cv2.imwrite(savepath + "/" + videoname +'_' + str(cnt) + '.jpg', clahe_img)
                update_json(savepath + "/" + videoname +'_' + str(cnt), videoname +'_' + str(cnt), water_info, h, w, Databin, videonamemp4, video_time, f)   
            else:
                newFrame = copy.deepcopy(image)
                sim = getSimilarity(Memory, newFrame)
                if sim <= threshold:
                    Databin = img_to_b64(image)
                    clahe_img = clahe_image(image)
                    h, w, _ = image.shape
                    if (h == 3840) and  (w == 2160):
                        clahe_img = cv2.rotate(clahe_img, cv2.ROTATE_90_CLOCKWISE)
                        h = 2160
                        w = 3840
                    video_time = str(datetime.timedelta(seconds= f/60))
                    cv2.imwrite(savepath + "/" + videoname +'_' + str(cnt) + '.jpg', clahe_img)
                    update_json(savepath + "/" + videoname +'_' + str(cnt), videoname +'_' + str(cnt), water_info, h, w, Databin, videonamemp4, video_time, f) 
                    
                    Memory = copy.deepcopy(image)
                    nFrame += 1
                else:
                    nFrame += 1
                    pass 
                       
        if cv2.waitKey(10) == 27:
            break    

def preprocess_img(image, savepath, water_info):
    imagenamejpg = os.path.split(image)[-1]
    imagename = os.path.splitext(imagenamejpg)[0]
    Databin = img_to_b64(image)
    image = cv2.imread(image)
    clahe_img = clahe_image(image)
    h, w, _ = image.shape
    if (h == 3840) and  (w == 2160):
        clahe_img = cv2.rotate(clahe_img, cv2.ROTATE_90_CLOCKWISE)
        h = 2160
        w = 3840
    elif (h not in [2160, 3840]) or (w not in [3840, 2160]):
        sg.Popup(imagename + '이미지 비율이 안 맞습니다.'+str(h) + ' ' + str(w), font =("Arial", 15), keep_on_top=True)
        return False, water_info 
    cv2.imwrite(savepath + "/" + imagenamejpg, clahe_img)
    update_json(savepath + "/" + imagename, imagename, water_info, h, w, Databin, None, None, None) 

def get_waterinfo(info_path):
    water_info_cols = ['Transparency', 'Latitude', 'Longitude', 'CDist', 'Site_Type', 'Depth']
    try:
        wb = openpyxl.load_workbook(info_path)
        # 수질 환경정보 excel 파일을 불러와 각 cell의 값을 list로 저장
        ws1 = wb['Sheet1']
        water_info = {}
    except:
        sg.Popup('excel파일이 경로에 없거나 Sheet 네임이 \n "Sheet1"으로 정확하게 입력되었는지 확인해주세요.', font =("Arial", 15), keep_on_top=True)
    
    for c in range(1, 6):
        if ws1.cell(1,c).value not in water_info_cols:
            sg.Popup('환경정보 파일의 열명이 Format에 맞지 않습니다. \nExcel 파일을 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
            return False, water_info
        if (ws1.cell(2,c).value == None) or (type(ws1.cell(2,c).value) == str):
            sg.Popup('환경정보 값이 누락되었거나 형식에 맞지 않습니다.\nExcel 파일을 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
            return False, water_info
        else:
            water_info[ws1.cell(1,c).value] = ws1.cell(2,c).value
    if ws1.cell(1, 6).value not in water_info_cols:
        sg.Popup('Excel 파일 내 Depth 열명이 올바른 지 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
        return False, water_info
    else:
        water_info[ws1.cell(1,6).value] = ws1.cell(2,6).value
    if (ws1.cell(2, 6).value == None) or (type(ws1.cell(2, 6).value) == float):
        sg.Popup('Depth 값이 누락되었거나 형식에 맞지 않습니다.\nExcel 파일을 확인해 주세요.', font =("Arial", 15), keep_on_top=True)
        return False, water_info   
    else:
        water_info[ws1.cell(1,6).value] = ws1.cell(2,6).value  
    return True, water_info

def check_waterinfo(water_info):
    flag_array = [True for i in range(1, 6)]
    if 0 > water_info['Transparency'] or water_info['Transparency'] > 15:
        flag_array[0] = False
    if 33.11 > water_info['Latitude'] or water_info['Latitude'] > 38.61:
        flag_array[1] = False
    if 124.6 > water_info['Longitude'] or water_info['Longitude'] > 131.87:
        flag_array[2] = False
    if water_info['Site_Type'] not in [1, 2]:
        flag_array[3] = False
    if water_info['Depth'] not in ['A', 'B', 'C']:
        flag_array[4] = False
    return flag_array

m = MakeGUI()
window = m.makegui()

while True:    
    event, values = window.read()     
    progress_bar = window['progress']
    if event == 'Run':
        Datapath = values['DataPath']
        info_path = values['WaterinfoPath']
        flag, water_info = get_waterinfo(info_path)
    
        if not flag:
            continue

        flag_array = check_waterinfo(water_info)
        if False in flag_array:
            sg.Popup('수질 환경정보 파일에 정상 범위를 벗어난 값이 있습니다.', keep_on_top=True)
            continue

        videolist = glob(Datapath + "/*.mp4")
        imagelist = glob(Datapath + "/*.jpg")
        datalength = len(videolist) + len(imagelist)
        progress_bar.UpdateBar(0, datalength)

        if len(videolist) > 0 :
            interval = int(values['intervals'])
        cnt = 1
        savepath = os.path.split(Datapath)[0] + '/' + str(os.path.split(Datapath)[1]) + '_Preprocessed'
        os.makedirs(savepath, exist_ok=True)
        for video in videolist:
            progress_bar.UpdateBar(cnt, datalength)
            videoname = os.path.split(video)[-1]
            preprocess_video(video, interval, savepath, water_info)
            cnt += 1
   
        for image in imagelist:
            progress_bar.UpdateBar(cnt, datalength)
            imagenamejpg = os.path.split(image)[-1]
            imagename = os.path.splitext(imagenamejpg)[0]
            preprocess_img(image, savepath, water_info)
            cnt += 1

        if cnt == datalength +1 :
            sg.Popup('정제 완료^^!', font =("Arial", 13), keep_on_top=True)
            break

    if event in (None, 'Exit'):
        break




