import PySimpleGUI as sg
from glob import glob
import os
import subprocess
import sys
import json
import openpyxl
import numpy as np
from glob import glob
from datetime import date

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')
        integrity = [
                  [sg.Text('INTEGRITY CHECK', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (10,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Text('EXCEL PATH', font =("Arial", 15)), sg.InputText(key='Excel', font =("Arial", 15), size = (4,1)), sg.FileBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1), key='excel')],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Labelme', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('FLAG', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Done', font =("Arial", 15, 'bold'), size=(25, 1))]
                    ]

        window = sg.Window('Integrity Check for PM', integrity, resizable=True, grab_anywhere = True, element_justification='c') 
        return window

def getCWid():
    userID = ''
    try:
        jsonfile = glob(values['Path'] + '/*.json')[0]
        with open(jsonfile) as jsontosave:
            objects = json.load(jsontosave)
            userID = objects['ID'].upper()
    except:
        sg.Popup("json파일이 없습니다." + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
    return userID

def runLabelme():
    conda_dirs = ["c:/Users/users/anaconda3", "d:/Users/users/anaconda3", 
                  "c:/Users/Admin/anaconda3", "d:/Users/Admin/anaconda3", "c:/Users/admin/anaconda3", "d:/Users/admin/anaconda3", 
                  "c:/Users/Administrator/anaconda3", "d:/Users/Administrator/anaconda3",
                  "c:/Programdata/anaconda3", "d:/Programdata/anaconda3", 
                  "c:/Users/{}/Anaconda3".format(os.getlogin()), "d:/Users/{}/Anaconda3".format(os.getlogin()), 
                  "c:/사용자/{}/Anaconda3".format(os.getlogin()), "d:/사용자/{}/Anaconda3".format(os.getlogin()),
                  "c:/anaconda3", "c:/Anaconda3", "d:/anaconda3", "d:/Anaconda3"]
    try:
        for Dir in conda_dirs:
            if os.path.isdir(Dir):
                python_file = Dir + "/envs/nia/pythonw.exe"
                labelme_file =  python_file[:-11] + "/Lib/site-packages/labelme/__main__.py"

            if Dir == "no dir":
                sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
        
        if np.__version__[:4] == '1.22':
        
            subprocess.Popen(python_file + ' ' + labelme_file,  stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)
            return 
    except:
        sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)

try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

try:
    flag_cnt = 0
    while True:    
        event, values = window.read()    
        if event == 'Labelme':
            runLabelme()
            userid = getCWid()
        
        if event =='FLAG':
            flag_cnt += 1

        if event == 'Done':
            sg.Popup("검수를 완료하셨습니다. \nExcel파일을 확인하시어 각 CW작업자분들에게 맞게 Feedback 부탁드립니다.", font =("Arial", 15), keep_on_top=True)   
            wb = openpyxl.load_workbook(values['excel'])
            ws = wb['Sheet2']
            try:
                for datecol in range(1,30):
                    if str(date.today()) == str(ws.cell(3, datecol).value)[:10]:
                        todaycol = datecol
            except:
                sg.Popup("Excel 파일 내에 오늘 날짜가 없습니다.", font =("Arial", 15), keep_on_top=True)   
            rowflag = False
            for row in range(4, 13):
                if userid == ws.cell(row, 1).value:
                    ws.cell(row, todaycol).value = flag_cnt
                    rowflag = True
            if rowflag == False:
                sg.Popup("엑셀 파일 안에 검수 대상 CW 아이디가 없습니다.", font =("Arial", 15), keep_on_top=True)   
            wb.save(values['excel'])
            flag_cnt = 0   

            break
        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()