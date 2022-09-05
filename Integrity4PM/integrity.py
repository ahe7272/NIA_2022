from operator import truediv
import PySimpleGUI as sg
from glob import glob
import os
import subprocess
import sys
import json
import openpyxl
import numpy as np
from datetime import date
import pandas as pd 
import shutil

class MakeGUI():
    def makegui(self):
        sg.theme('DarkAmber')
        integrity = [
                  [sg.Text('INTEGRITY CHECK', font =("Arial", 30, 'bold'), text_color = 'Skyblue')],
                  [sg.Text('PATH', font =("Arial", 15)), sg.InputText(key='Path', font =("Arial", 15), size = (10,1)), sg.FolderBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Text('EXCEL PATH', font =("Arial", 15)), sg.InputText(key='Excel', font =("Arial", 15), size = (4,1)), sg.FileBrowse('SELECT',font =("Arial", 15, 'bold') , size=(10, 1), key='excel')],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Text('검수 주차', font =("Arial", 15, 'bold'), size=(8, 1)), sg.InputText(key='Week', font =("Arial", 15),  size = (10,1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Labelme', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],
                  [sg.Button('Image name list to Excel', font =("Arial", 15, 'bold'), size=(25, 1))],
                  [sg.Text('', font =("Arial", 10, 'bold'))],            
                  [sg.Button('Done', font =("Arial", 15, 'bold'), size=(25, 1))]
                    ]

        window = sg.Window('Integrity Check for PM', integrity, resizable=True, grab_anywhere = True, element_justification='c') 
        return window

def list2excel(userid):
    all_filenames = []
    try:
        for (path, dir, files) in os.walk(values['Path']):
            for jpg in files:
                if jpg[-4:] == '.jpg':
                    all_filenames += [[jpg]]
        listdf = pd.DataFrame.from_records(all_filenames)
        filename = values['Path'] + '/' + userid + '_' + str(date.today()) + '0.xlsx'
        try: 
            if os.path.exists(filename): 
                lastfileno = max([int(x.split(str(date.today()))[-1][0]) for x in glob(filename[:-7]+'*.xlsx')])
                filename = values['Path'] + '/' + userid + '_' + str(date.today()) + str(lastfileno + 1) + '.xlsx'
                listdf.to_excel(filename)
                listlen = len(all_filenames)
            else:
                listdf.to_excel(filename)
                listlen = len(all_filenames)

        except OSError: 
            print("directory를 생성할 수 없습니다.")
            exit()

        return filename, listlen
    except:
        sg.Popup("파일명리스트 엑셀파일 생성에 실패했습니다." + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
    

def getCWid():
    userID = ''
    try:
        jsonfile = glob(values['Path'] + '/*.json')[0]
        with open(jsonfile) as jsontosave:
            objects = json.load(jsontosave)
            userID = objects['ID'].upper()
    except:
        sg.Popup("json파일이 없습니다." + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
    return userID

def runLabelme():
    conda_dirs = ["c:/Users/users/anaconda3", "d:/Users/users/anaconda3", 
                  "c:/Users/Admin/anaconda3", "d:/Users/Admin/anaconda3", "c:/Users/admin/anaconda3", "d:/Users/admin/anaconda3", 
                  "c:/Users/Administrator/anaconda3", "d:/Users/Administrator/anaconda3",
                  "c:/Programdata/anaconda3", "d:/Programdata/anaconda3", 
                  "c:/Users/{}/Anaconda3".format(os.getlogin()), "d:/Users/{}/Anaconda3".format(os.getlogin()), 
                  "c:/사용자/{}/Anaconda3".format(os.getlogin()), "d:/사용자/{}/Anaconda3".format(os.getlogin()),
                  "c:/anaconda3", "c:/Anaconda3", "d:/anaconda3", "d:/Anaconda3"]
    try:
        for Dir in conda_dirs:
            if os.path.isdir(Dir):
                python_file = Dir + "/envs/nia/pythonw.exe"
                labelme_file =  python_file[:-11] + "/Lib/site-packages/labelme/__main__.py"

            if Dir == "no dir":
                sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)
        
        if np.__version__[:4] == '1.22':
        
            subprocess.Popen(python_file + ' ' + labelme_file,  stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)
            return 
    except:
        sg.Popup("anaconda 경로를 찾지 못했습니다. 담당자에게 알려주세요. 아래 메세지를 알려주세요.\n\n" + subprocess.getoutput('where labelme'), font =("Arial", 15), keep_on_top=True)

try:
    os.chdir(sys._MEIPASS)
    print(sys._MEIPASS)
except:
    os.chdir(os.getcwd())

m = MakeGUI()
window = m.makegui()

try:
    flag_cnt = 0
    while True:    
        event, values = window.read()    
        if event == 'Labelme':
            runLabelme()
            userid = getCWid()
        if event == 'Image name list to Excel':
            listexcelpath, listlen = list2excel(userid)

        if event == 'Done':
            sg.Popup("검수를 완료하셨습니다.", font =("Arial", 15), keep_on_top=True)   
            PASS_path = values['Path'] + '/' + userid + '_PASS' + '/'
            FAIL_path = values['Path'] + '/' + userid + '_FAIL' + '/'
            os.makedirs(PASS_path, exist_ok=True)
            os.makedirs(FAIL_path, exist_ok=True)

            listwb = openpyxl.load_workbook(listexcelpath)
            listws = listwb['Sheet1']
            try: 
                for passROW in range(2, listlen+2):
                    Pure_filename = listws.cell(passROW,2).value[:-4]
                    if listws.cell(passROW,3).value == 0 :                    
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', PASS_path)
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', PASS_path)
                    elif (listws.cell(passROW,3).value != None) and (listws.cell(passROW,3).value > 0): 
                        flag_cnt += listws.cell(passROW,3).value
                        shutil.move(values['Path'] + '/' + Pure_filename +'.jpg', FAIL_path)
                        shutil.move(values['Path'] + '/' + Pure_filename +'.json', FAIL_path)
                    else:
                        continue
            except:
                sg.Popup("반려, 통과 파일을 옮기는데 실패했습니다.", font =("Arial", 15), keep_on_top=True)   
            wb = openpyxl.load_workbook(values['excel'])
            ws = wb['Sheet2']

            try:
                for datecol in range(3,8):
                    if ws.cell(3, datecol).value == None:
                        continue
                    elif ws.cell(3, datecol).value == values['Week']:
                        todaycol = datecol
            except:
                sg.Popup("Excel 파일 내에 검수할 주차 이름이 없습니다.", font =("Arial", 15), keep_on_top=True)   
            for row in range(4, 12):
                if ws.cell(row, 1).value == None:
                    continue
                if userid == ws.cell(row, 1).value:
                    ws.cell(row, todaycol).value += flag_cnt
                  
            wb.save(values['excel'])
            flag_cnt = 0   
            break
        if event == sg.WIN_CLOSED or event in (None, 'Exit'):
            break
        if event in (None, 'Exit'):
            break

except Exception as e:
   print(e)
window.close()