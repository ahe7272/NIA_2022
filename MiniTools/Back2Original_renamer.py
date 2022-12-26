import json
import os 
import pandas as pd
import openpyxl

# 파일명이 bbox_00001 형태로 바뀐 원본 사진을 원래 이전 이름 (fixed_... 형식)으로 바꾸는 작업
running_path = input('경로: ')
excel_path = input('log 엑셀 경로: ')
ori_path = input('원본 경로: ')

wb = openpyxl.load_workbook(excel_path)
file_list = wb['Name_change_log_Bbox']

old_names = {}
for (path, dir, files) in os.walk(running_path): 
    for item in files:
        if item[-4:] == 'json':
            old_names[item.split('.')[0]] = path + '/' 

# for (path, dir, files) in os.walk(running_path): 
#     for item in files:
#         if item[-4:] == '.jpg':
#             old_names[item.split('.')[0]] = path + '/' 


def getjson(jsonfile):
    with open(jsonfile) as j:
        objects = json.load(j)        
    return objects

# log 엑셀에서 old_names의 키로 저장된 값들이 있는 경우 해당 행의 1열값(바꾸기 전 원래 이름)으로 해당 이미지명을 바꾸는 작업  
new_names = []
for row in range(1, file_list.max_row+1):
    if file_list.cell(row, 2).value.split('.')[0] in old_names.keys():
        old_jpg = old_names[file_list.cell(row, 2).value.split('.')[0]] + file_list.cell(row, 2).value.split('.')[0] + '.jpg'
        new_jpg = old_names[file_list.cell(row, 2).value.split('.')[0]] + file_list.cell(row, 1).value.split('.')[0] + '.jpg'
        old_json = old_names[file_list.cell(row, 2).value.split('.')[0]] + file_list.cell(row, 2).value.split('.')[0] + '.json'
        new_json =  old_names[file_list.cell(row, 2).value.split('.')[0]] + file_list.cell(row, 1).value.split('.')[0] + '.json'
        old_original = ori_path + '/' + file_list.cell(row, 2).value.split('.')[0] + '.jpg'
        # print(old_original)
        new_original = ori_path + '/' + '_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0] + '.jpg'
        # print(old_original)
        # print(new_original)
        os.rename(old_jpg, new_jpg)
        os.rename(old_json, new_json)
        # os.rename(old_original, new_original)