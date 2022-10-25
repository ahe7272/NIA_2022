import json
import os 
import shutil
import openpyxl

excel_path = 'C:/Users/Administrator/Desktop/Test/Name_change_log_Polygon.xlsx'
new_original_path = 'C:/Users/Administrator/Desktop/Test/new_ori'
old_original_path = 'C:/Users/Administrator/Desktop/Test/old_ori'
save_path = 'C:/Users/Administrator/Desktop/Test/to'
removed_path = 'C:/Users/Administrator/Desktop/Test/removed'

wb = openpyxl.load_workbook(excel_path)
ws1 = wb['Sheet1']

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
        Jsonfile.close()
    return objects

old_list = []
for (path, dir, files) in os.walk(old_original_path): 
    old_list += [path + '/' + Jpg for Jpg in files if Jpg.lower().endswith('.jpg')]

current_picture = []
for row in range(2, ws1.max_row+1):
    for old_picture in old_list:
        if '_'.join(ws1.cell(row,1).value.split('_')[4:]) == '_'.join((old_picture.split('/')[-1]).split('_')[2:4]) +'.jpg':
            current_picture += [old_picture]
    if len(current_picture) > 1:
        print(ws1.cell(row,1).value +'에 대해 ' + str(current_picture) + '의 사진이 존재합니다. \n')
        which_one = input('어떤 사진을 사용할까요? (0, 1, 2 ...)')
        try: 
            video_name = ws1.cell(row,1).value.split('_')[4]
            fno_jpg = ws1.cell(row,1).value.split('_')[5]
            original_path = new_original_path + '/' + video_name + '_' + fno_jpg.split('.jpg')[0].zfill(3) +'.jpg'
            new_name_path = save_path + '/' + ws1.cell(row,2).value
            os.rename(original_path, new_name_path)
            shutil.move(current_picture[int(which_one)], removed_path +'/' + current_picture[int(which_one)].split('/')[-1])

            current_picture = []
            continue
        except:
            break
    elif len(current_picture) == 1:
        video_name = ws1.cell(row,1).value.split('_')[4]
        fno_jpg = ws1.cell(row,1).value.split('_')[5]
        original_path = new_original_path + '/' + video_name + '_' + fno_jpg.split('.jpg')[0].zfill(3) +'.jpg'
        new_name_path = save_path + '/' + ws1.cell(row,2).value
        os.rename(original_path, new_name_path)
        current_picture = []