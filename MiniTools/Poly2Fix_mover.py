import os 
import openpyxl
import shutil
import json

excel_path = input('파일명 리스트 엑셀 파일 경로를 입력해주세요. ')
to_path = input('저장할 폴더 경로를 입력해주세요. ')
from_path = input('가져올 파일이 있는 폴더 경로를 입력해주세요. ')

wb = openpyxl.load_workbook(excel_path)
file_list = wb['Sheet1']

file_names = []
for row in range(1, file_list.max_row+1):
    file_names += [file_list.cell(row,2).value]

for (path, dir, files) in os.walk(from_path): 
    for item in files:
        if item[-4:] == '.jpg':
            if item in file_names:
                shutil.copy(path + '/' + item, to_path + '/' + item)
                shutil.copy(path + '/' + item[:-4] + '.json', to_path + '/' + item[:-4] + '.json')
                