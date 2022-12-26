import json
import os 
import pandas as pd
import openpyxl

running_path = input('경로: ')
excel_path = input('log 엑셀 경로: ')
save_path = input('저장 경로: ') + '/'
wb = openpyxl.load_workbook(excel_path)
file_list = wb['Name_change_log_Polygon']

old_names = {}
for (path, dir, files) in os.walk(running_path): 
    for item in files:
        if item[-4:] == 'json':
            old_names[item.split('.')[0]] = path + '/' 



def getjson(jsonfile):
    with open(jsonfile) as j:
        objects = json.load(j)        
    return objects


new_names = []
for row in range(1, file_list.max_row+1):
    # print('_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0])
    if file_list.cell(row, 1).value.split('.')[0] in old_names.keys():
        old_json = old_names[file_list.cell(row, 1).value.split('.')[0]] + file_list.cell(row, 1).value.split('.')[0]+ '.json'
        new_json = save_path + file_list.cell(row, 2).value.split('.')[0] + '.json'
        old_objects = getjson(old_json)
        new_objects = getjson(new_json)
        new_objects['shapes'] = old_objects['shapes']
        with open(new_json, 'w') as j:
            json.dump(new_objects, j, indent='\t')
            j.close()
