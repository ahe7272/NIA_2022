import json
import os 
import pandas as pd
import openpyxl

image_path = input('경로: ')
Korean = input('한글명: ')
English = input('영문명: ')

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
        Jsonfile.close()
    return objects

for (root, dir, files) in os.walk(image_path): 
    for item in files:
        if item[-4:] == 'xlsx':
            wb = openpyxl.load_workbook(root + '/' + item)
            wbpath = root + '/' + item
            ws1 = wb['Sheet1']
            ws1.cell(2, 5).value = ws1.cell(2,13).value + (float(str(ws1.cell(2,14).value).split('.')[0])/60) + (float(str(ws1.cell(2,14).value).split('.')[0])/3600)
            ws1.cell(2, 6).value = ws1.cell(2,15).value + (float(str(ws1.cell(2,16).value).split('.')[0])/60) + (float(str(ws1.cell(2,16).value).split('.')[0])/3600)
            ws2 = wb['Sheet2']
            continue
        
        new_name = ''.join(item.split(' '))
        if Korean != '':
            new_name = English.join(new_name.split(Korean))
        old = root + '/' + item
        new = root + '/' + new_name
        os.rename(old, new)

for row in range(1, ws2.max_row+1):
    if (ws2.cell(row, 1).value) == None:
        print('Name 행의 끝이 ' + str(row) +'이 맞나요?')
        continue
    new_image_name = ''.join(ws2.cell(row, 1).value.split(' '))
    if Korean != '':
        new_image_name = English.join(new_image_name.split(Korean))
    ws2.cell(row, 1).value = new_image_name
    if ws2.cell(row, 2).value == 'Asterias Amurensis':
        ws2.cell(row, 2).value = 'Asterias_amurensis'
    elif ws2.cell(row, 2).value == 'Asterina Pectinifera':
        ws2.cell(row, 2).value = 'Asterina_pectinifera'
    elif ws2.cell(row, 2).value == 'EckloniaCava':
        ws2.cell(row, 2).value = 'Ecklonia_cava'
    elif ws2.cell(row, 2).value == 'Heliocidaris Crassispina':
        ws2.cell(row, 2).value = 'Heliocidaris_crassispina'
    elif ws2.cell(row, 2).value == 'SeaHare':
        ws2.cell(row, 2).value = 'Sea_hare'
    elif ws2.cell(row, 2).value == 'Turbo Cornutus':
        ws2.cell(row, 2).value = 'Turbo_cornutus'
    
wb.save(wbpath)

# # 엑셀 내 이름 바꾸기
# wb = openpyxl.load_workbook('D:/34th_S_W_image/0830_경남 통영_이미지_1-3.xlsx')
# ws1 = wb['Sheet1']
# ws1.cell(2, 5).value = ws1.cell(2,13).value + (float(str(ws1.cell(2,14).value).split('.')[0])/60) + (float(str(ws1.cell(2,14).value).split('.')[0])/3600)
# ws1.cell(2, 6).value = ws1.cell(2,15).value + (float(str(ws1.cell(2,16).value).split('.')[0])/60) + (float(str(ws1.cell(2,16).value).split('.')[0])/3600)
# ws2 = wb['Sheet2']
# for row in range(1, ws2.max_row+1):
#     if (ws2.cell(row, 1).value) == None:
#         print('Name 행의 끝이 ' + str(row) +'이 맞나요?')
#         continue
#     new_image_name = ''.join(ws2.cell(row, 1).value.split(' '))
#     if Korean != '':
#         new_image_name = English.join(new_image_name.split(Korean))
#     ws2.cell(row, 1).value = new_image_name
#     if ws2.cell(row, 2).value == 'Asterias Amurensis':
#         ws2.cell(row, 2).value = 'Asterias_amurensis'
#     elif ws2.cell(row, 2).value == 'Asterina Pectinifera':
#         ws2.cell(row, 2).value = 'Asterina_pectinifera'
#     elif ws2.cell(row, 2).value == 'EckloniaCava':
#         ws2.cell(row, 2).value = 'Ecklonia_cava'
#     elif ws2.cell(row, 2).value == 'Heliocidaris Crassispina':
#         ws2.cell(row, 2).value = 'Heliocidaris_crassispina'
#     elif ws2.cell(row, 2).value == 'SeaHare':
#         ws2.cell(row, 2).value = 'Sea_hare'
#     elif ws2.cell(row, 2).value == 'Turbo Cornutus':
#         ws2.cell(row, 2).value = 'Turbo_cornutus'

# wb.save('D:/34th_S_W_image/0830_경남 통영_이미지_1-3.xlsx')

# # # 사진 이름 바꾸기
# for (root, dir, files) in os.walk('D:/33rd_S_W_image/20220820/1-1'): 
#     for item in files:
#         if item[-4:] == 'xlsx':
#             continue
#         # new_name = ''.join(item.split(' '))
#         # if Korean != '':
#             # new_name = English.join(new_name.split(Korean))
#         new_name = item[:8] + '_' + item[8:] 
#         print(new_name)
#         old = root + '/' + item
#         new = root + '/' + new_name
#         os.rename(old, new)

# s2wb = openpyxl.load_workbook('D:/23rd/20220929/1-1/0929_울릉군 독도_1-1.xlsx')
# ws_copy = s2wb.copy_worksheet(s2wb['Sheet2'])

# for (root, dir, files) in os.walk(image_path): 
#     for item in files:
#         if item[-4:] == 'xlsx':
#             wb = openpyxl.load_workbook(root + '/' + item)
#             wbpath = root + '/' + item
#             ws1 = wb['Sheet1']
#             ws1.cell(2, 5).value = ws1.cell(2,13).value + (float(str(ws1.cell(2,14).value).split('.')[0])/60) + (float(str(ws1.cell(2,14).value).split('.')[0])/3600)
#             ws1.cell(2, 6).value = ws1.cell(2,15).value + (float(str(ws1.cell(2,16).value).split('.')[0])/60) + (float(str(ws1.cell(2,16).value).split('.')[0])/3600)
#             # print(ws1.cell(2, 5).value)
#             # wb.create_sheet('Sheet2',1)
#             # ws2 = wb['Sheet2']
#             # ws2 = ws_copy
        
#             wb.save(wbpath)