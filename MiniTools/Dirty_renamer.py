import json
import os 
import pandas as pd
import openpyxl

running_path = input('경로: ')
excel_path = input('log 엑셀 경로: ')

wb = openpyxl.load_workbook(excel_path)
file_list = wb['24th_Name_change_log_b']

old_names = {}
for (path, dir, files) in os.walk(running_path): 
    for item in files:
        if item[-4:] == '.jpg':
            old_names[item.split('.')[0]] = path + '/' 

new_names = []
for row in range(1, file_list.max_row+1):
    # print('_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0])
    if file_list.cell(row, 1).value.split('.')[0] in old_names.keys():
        old_jpg = old_names[file_list.cell(row, 1).value.split('.')[0]] + file_list.cell(row, 1).value.split('.')[0] + '.jpg'
        new_jpg = old_names[file_list.cell(row, 1).value.split('.')[0]] + file_list.cell(row, 2).value.split('.')[0] + '.jpg'
        old_json = old_names[file_list.cell(row, 1).value.split('.')[0]] + file_list.cell(row, 1).value.split('.')[0] + '.json'
        new_json = old_names[file_list.cell(row, 1).value.split('.')[0]] + file_list.cell(row, 2).value.split('.')[0] + '.json'
        print(old_jpg)
        print(new_jpg)
        os.rename(old_jpg, new_jpg)
        os.rename(old_json, new_json)
