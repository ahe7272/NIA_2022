import json
import os 
import pandas as pd
import openpyxl

# 파일명이 bbox_00001 형태로 바뀐 원본 사진을 원래 이전 이름 (fixed_... 형식)으로 바꾸는 작업
running_path = input('경로: ')
excel_path = input('log 엑셀 경로: ')
original_path = input('원본 경로: ')
wb = openpyxl.load_workbook(excel_path)
file_list = wb['Sheet1']

# 경로를 검색하면서 jpg로 끝나는 파일의 이름을 old_names Dictionary에 저장
old_names = {}
for (path, dir, files) in os.walk(running_path): 
    for item in files:
        if item[-4:] == '.jpg':
            old_names['_'.join(item.split('_')[2:]).split('.')[0]] = path + '/' 

# log 엑셀에서 old_names의 키로 저장된 값들이 있는 경우 해당 행의 1열값(바꾸기 전 원래 이름)으로 해당 이미지명을 바꾸는 작업  
new_names = []
for row in range(1, file_list.max_row+1):
    # print('_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0])
    if '_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0] in old_names.keys():
        old_path = old_names['_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0]] + '_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0]+ '.jpg'
        new_name = old_names['_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0]] + file_list.cell(row,2).value.split('.')[0] + '.jpg'
        new_origianl = original_path + '/' + file_list.cell(row,2).value.split('.')[0] + '.jpg'
        old_original = original_path + '/' + '_'.join(file_list.cell(row, 1).value.split('_')[2:]).split('.')[0]+ '.jpg'
        print(old_original)
        os.rename(new_origianl, old_original)