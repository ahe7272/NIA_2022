import json
import os
import pandas as pd
import cv2 
import openpyxl
import copy

path = input("비디오와 야장이 있는 경로를 입력하세요.\n")

def getjson(jsonfile):
    with open(jsonfile) as j:
        objects = json.load(j)
    return objects

meta_dict = {'Latitude':0, 'Longitude':0, 'Video_length': 0}

df_all = pd.DataFrame.from_dict(list(copy.deepcopy(meta_dict).keys())).transpose()

for root, dirs, files in os.walk(path):
    try:
        excel = [Excel for Excel in files if Excel.lower().endswith('xlsx')][0]
        videos = [Video for Video in files if Video.lower().endswith('mp4')]
    except:
        continue
    duration = 0
    wb = openpyxl.load_workbook(root + '/' + excel)
    ws1 = wb['Sheet1']
    try:
        meta_dict['Latitude'] = ws1.cell(2, 8).value + (ws1.cell(2, 9).value/60)
        meta_dict['Longitude'] =  ws1.cell(2, 10).value + (ws1.cell(2,11).value/60)
    except:
        print(root)
        continue

    for Video in videos:
        video = cv2.VideoCapture(root + '/' + Video)
        fps = video.get(cv2.CAP_PROP_FPS)      # OpenCV v2.x used "CV_CAP_PROP_FPS"
        frame_count = int(video.get(cv2.CAP_PROP_FRAME_COUNT))
        duration += frame_count/fps
    meta_dict['Video_length'] = duration

    df_single = pd.DataFrame(list(meta_dict.values())).transpose()    
    df_all = pd.concat([df_all, df_single])
    print(df_single)

df_all.to_excel('C:/Users/Administrator/Downloads/Total.xlsx')