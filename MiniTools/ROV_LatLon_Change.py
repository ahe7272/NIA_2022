import json
import os 
import shutil
import openpyxl

excel_path = input('파일명 리스트 엑셀 파일 경로를 입력해주세요. ')
to_path = input('속성을 변경할 파일이 있는 폴더 경로를 입력해주세요. ')
# from_path = input('속성을 가져올 파일이 있는 폴더 경로를 입력해주세요. ')

wb = openpyxl.load_workbook(excel_path)
ws1 = wb['Sheet1']

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
        Jsonfile.close()
    return objects

data = {}
for row in range(1, ws1.max_row):
    try:
        data[ws1.cell(row,4).value] = [ws1.cell(row,5).value, ws1.cell(row,6).value] 
    except:
        continue

for (path, dir, files) in os.walk(to_path): 
    for item in files:
        if item[-5:] == '.json':
            if ('_'.join(item.split('_')[2:-1])) in data.keys():
                objects = getjson(to_path + '/' + item)
                objects['Longitude'] = data['_'.join(item.split('_')[2:-1])][0]
                objects['Latitude'] = data['_'.join(item.split('_')[2:-1])][1]
                # print(objects['Longitude'], objects['Latitude'])
    #     to_objects = getjson(to_path + '/' + ws1.cell(row,2).value[:-4] + '.json')
    #     to_objects['Longitude'] = from_objects['Longitude'] 
    #     to_objects['Latitude'] = from_objects['Latitude']
    #     print('From: ', from_objects['Longitude'], from_objects['Latitude'])
    #     print('To: ', to_objects['Longitude'], to_objects['Latitude'])
                with open(to_path + '/' + item,  'w') as j:
                    json.dump(objects, j, indent='\t')
                    j.close()

# addfrom_path = 'D:/1st_Original/Bbox'
# addto_path = 'D:/1st_Original/Test/Polygon'

# for (addfrompath, dir, addfromfiles) in os.walk(addfrom_path): 
#     for addfromitem in addfromfiles:
#         if addfromitem[-5:] == '.json':
#             if os.path.isfile(addto_path + '/' + addfromitem):
#                 shutil.copy(addfrom_path +'/' + addfromitem[:-5] +'_original.jpg', addto_path + '/' + addfromitem[:-5] +'_original.jpg')  
            
#                 from_objects = getjson(addfrompath + '/' + addfromitem)
#                 to_objects = getjson(addto_path + '/' + addfromitem)
#                 to_objects['Source_video'] = from_objects['Source_video'] 
#                 to_objects['Video_time'] = from_objects['Video_time'] 
#                 to_objects['Frame_no'] = from_objects['Frame_no'] 
#                 to_objects['Collection_method'] = from_objects['Collection_method'] 
#                 to_objects['Distance'] = 0.5
#                 with open(addto_path + '/' + addfromitem, 'w') as j:
#                     json.dump(to_objects, j, indent='\t')
#                     j.close()