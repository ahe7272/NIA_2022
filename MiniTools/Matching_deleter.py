import os
import shutil
import openpyxl

excel_path = input('삭제리스트 엑셀 파일 경로를 입력해주세요. ')
remove_path = input('삭제할 파일이 있는 폴더 경로를 입력해주세요. ')
removed_path = 'C:/Users/Administrator/Desktop/removed'
wb = openpyxl.load_workbook(excel_path)
ws1 = wb['Sheet1']

for row in range(2, ws1.max_row+1):
    if os.path.isfile(remove_path + '/' + ws1.cell(row,1).value):
        shutil.move(remove_path + '/' + ws1.cell(row,1).value, removed_path + '/' + ws1.cell(row,1).value)  
        # shutil.move(remove_path + '/' + ws1.cell(row,1).value[:-4] + '.json', removed_path + '/' + ws1.cell(row,1).value[:-4] + '.json')  
        # shutil.move(remove_path + ' original/' + ws1.cell(row,1).value[:-4] + '_original.jpg', removed_path + '/' + ws1.cell(row,1).value[:-4] + '_original.jpg')  