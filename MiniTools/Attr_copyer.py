import json
import os 
import shutil
import openpyxl

excel_path = input('파일명 리스트 엑셀 파일 경로를 입력해주세요. ')
to_path = input('속성을 변경할 파일이 있는 폴더 경로를 입력해주세요. ')
from_path = input('속성을 가져올 파일이 있는 폴더 경로를 입력해주세요. ')

wb = openpyxl.load_workbook(excel_path)
ws1 = wb['Sheet1']

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
        Jsonfile.close()
    return objects

for row in range(2, ws1.max_row+1):
    # print(from_path + '/'  + ws1.cell(row,1).value[:-4] + '.json')
    # break
    # if os.path.isfile(from_path + '/' + ws1.cell(row,1).value[:-4] + '.json'):
        from_objects = getjson(from_path + '/' + ws1.cell(row,1).value[:-4] + '.json')
        to_objects = getjson(to_path + '/' + ws1.cell(row,2).value[:-4] + '.json')
        to_objects['Longitude'] = from_objects['Longitude'] 
        to_objects['Latitude'] = from_objects['Latitude']
        print('From: ', from_objects['Longitude'], from_objects['Latitude'])
        print('To: ', to_objects['Longitude'], to_objects['Latitude'])
        # with open(to_path + '/' + + ws1.cell(row,2).value, 'w') as j:
        #     json.dump(to_objects, j, indent='\t')
        #     j.close()

# addfrom_path = 'D:/1st_Original/Bbox'
# addto_path = 'D:/1st_Original/Test/Polygon'

# for (addfrompath, dir, addfromfiles) in os.walk(addfrom_path): 
#     for addfromitem in addfromfiles:
#         if addfromitem[-5:] == '.json':
#             if os.path.isfile(addto_path + '/' + addfromitem):
#                 shutil.copy(addfrom_path +'/' + addfromitem[:-5] +'_original.jpg', addto_path + '/' + addfromitem[:-5] +'_original.jpg')  
            
#                 from_objects = getjson(addfrompath + '/' + addfromitem)
#                 to_objects = getjson(addto_path + '/' + addfromitem)
#                 to_objects['Source_video'] = from_objects['Source_video'] 
#                 to_objects['Video_time'] = from_objects['Video_time'] 
#                 to_objects['Frame_no'] = from_objects['Frame_no'] 
#                 to_objects['Collection_method'] = from_objects['Collection_method'] 
#                 to_objects['Distance'] = 0.5
#                 with open(addto_path + '/' + addfromitem, 'w') as j:
#                     json.dump(to_objects, j, indent='\t')
#                     j.close()