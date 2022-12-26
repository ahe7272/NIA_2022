import json
import os 
import shutil
import openpyxl

match_path = '//nia/올빅뎃_조식동물/품질검수/18차(2022-11-30)/Polygon2Fix'
done_path = '//nia/올빅뎃_조식동물/품질검수/18차(2022-11-30)/Used'
excel_path = input('log 엑셀 경로: ')
wb = openpyxl.load_workbook(excel_path)
file_list = wb['Name_change_log_Polygon']


old_names = {}
for (path, dir, files) in os.walk(match_path): 
    for item in files:
        if item[-4:] == 'json':
            old_names[item.split('.')[0]] = path + '/' 

# log 엑셀에서 old_names의 키로 저장된 값들이 있는 경우 해당 행의 1열값(바꾸기 전 원래 이름)으로 해당 이미지명을 바꾸는 작업  
new_names = []
for row in range(1, file_list.max_row+1):
    if file_list.cell(row, 1).value.split('.')[0][20:] in old_names.keys():
        print(file_list.cell(row, 1).value.split('.')[0][20:])
        shutil.move(match_path + '/' + file_list.cell(row, 1).value.split('.')[0][20:] + '.jpg', done_path + '/' + file_list.cell(row, 1).value.split('.')[0][20:] + '.jpg')    
        shutil.move(match_path + '/' + file_list.cell(row, 1).value.split('.')[0][20:] + '.json', done_path + '/' + file_list.cell(row, 1).value.split('.')[0][20:] + '.json')    

 