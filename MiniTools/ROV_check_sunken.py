import os
import openpyxl
import pandas as pd 
import json

def getjson(jsonfile):
    with open(jsonfile) as Jsonfile:
        objects = json.load(Jsonfile)
        Jsonfile.close()
    return objects

path_input = input('경로: ')
excel_input = input('ROV 엑셀 경로: ')

wb = openpyxl.load_workbook(excel_input)
ws = wb['Sheet1']
ROV_dict = {}
for r in range(1, ws.max_row+1):
    ROV_dict[ws.cell(r,4).value] = [ws.cell(r,5).value, ws.cell(r,6).value]

for (path, dir, files) in os.walk(path_input): 
    if path.split('\\')[-1] == 'Originals':
        continue
    for item in files:
        if item[-4:] == 'json':
            objects = getjson(path + '/' + item)
            video_name = '_'.join(item.split('_')[2:-1])
            if video_name in ROV_dict.keys():
                objects['Longitude'] = ROV_dict[video_name][0]
                objects['Latitude'] = ROV_dict[video_name][1]
                with open(path + '/' + item, 'w') as j:
                    json.dump(objects, j, indent='\t')
                    j.close()

