import json
import os
from classes import debris_dict
import openpyxl
import copy
from natsort import natsorted


folder_path = input("폴더 경로를 입력하세요.\n")
classes_dict= debris_dict()

wb = openpyxl.load_workbook(input("엑셀 파일 경로를 입력하세요.\n"))
ws1 = wb['Sheet1']
target_videos = []
json_list = []

def getjson(jsonfile):
    with open(jsonfile) as j:
        objects = json.load(j)
    return objects

total = copy.deepcopy(classes_dict)
continuing = []

for row in range(1, 468):
    if ws1.cell(row, 1).value != None:
        target_videos += [ws1.cell(row, 1).value]
    else:
        for video in target_videos:
            ordered_jsons = []
            for path, dirs, files in os.walk(folder_path):
                json_list += [path + '/' + Json for Json in files if video in Json.split('_')]     
            json_list = natsorted(json_list)
            for target_json in json_list:
                objects = getjson(target_json)
                if len(objects['shapes']) == 0:
                    print(target_json)
                    continue  

                label_list = [label['label'] for label in objects['shapes']]     
                for labelname in label_list:
                    if labelname not in continuing:
                        continuing += [labelname]
                complement = list(set(continuing).difference(label_list))
                if complement:
                    for item in complement:
                        continuing.pop(continuing.index(item))
                        total[item] += 1

            json_list = []
        target_videos = []
print(total)


 
            
            
                
                    
# df = pd.DataFrame(list(total.items()), columns=['classname', 'Count'])
# df.to_excel(path + '/ObtainmentRate.xlsx')